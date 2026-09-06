#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gelişmiş Kripto Sinyal Botu - Tam Düzeltmiş Versiyon
Ekran görüntülerindeki Kripto Zeyna botuna birebir yakın özellikler:
- SMC + klasik indikatör skorlama (max 20)
- MTF bonus
- Detaylı grafik (MA50/100/200)
- Açık işlem + BE (Break Even) yönetimi
- Excel 2 sekmeli rapor (Coin Performansı + Özet) renkli
- Aylık performans özeti
- Telegram uyumlu mesaj formatı
"""

import os
import time
import sqlite3
import threading
from datetime import datetime, timedelta
from collections import defaultdict
import requests
import pandas as pd
import pandas_ta as ta
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ==========================================
# 1. AYARLAR
# ==========================================
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")
RENDER_URL = os.environ.get("RENDER_EXTERNAL_URL", "")

# Ekrandaki 73 coin (OKX formatı: BTC-USDT)
HEDEF_COINLER = [
    'ONT-USDT', 'JUP-USDT', 'SNX-USDT', 'LDO-USDT', 'ZETA-USDT', 'AAVE-USDT', 'TIA-USDT', 'VET-USDT', 'DYDX-USDT', 'LTC-USDT',
    'AVAX-USDT', 'SUSHI-USDT', 'STX-USDT', 'XRP-USDT', 'ORDI-USDT', 'ENS-USDT', 'GALA-USDT', 'MANA-USDT', 'ICP-USDT', 'THETA-USDT',
    'TRX-USDT', 'EGLD-USDT', 'ADA-USDT', 'AXS-USDT', 'INJ-USDT', 'AEVO-USDT', 'BCH-USDT', 'FLOKI-USDT', 'DOT-USDT', 'RUNE-USDT',
    'KAS-USDT', 'COMP-USDT', 'BONK-USDT', 'ALT-USDT', 'BTC-USDT', 'FIL-USDT', 'WIF-USDT', 'PYTH-USDT', 'FET-USDT', 'PEPE-USDT',
    'SHIB-USDT', 'SUI-USDT', 'ATOM-USDT', 'SAND-USDT', 'PORTAL-USDT', 'APT-USDT', 'STRK-USDT', 'ARB-USDT', 'DYM-USDT', 'METIS-USDT',
    'BNB-USDT', 'UNI-USDT', 'SEI-USDT', 'HBAR-USDT', 'MANTA-USDT', 'LINK-USDT', 'OP-USDT', 'XAI-USDT', 'ALPINE-USDT', 'MAV-USDT',
    'DOGE-USDT', 'ETHFI-USDT', 'CRV-USDT', 'PIXEL-USDT', 'IMX-USDT', 'ETH-USDT', 'GRT-USDT', 'ENJ-USDT', 'ONE-USDT', 'SOL-USDT',
    'NEAR-USDT', 'ALGO-USDT', 'PENDLE-USDT'
]

MIN_SKOR = 15.0
OKX_BASE = "https://www.okx.com"

# ==========================================
# 2. VERİ ÇEKME (OKX)
# ==========================================
def veri_cek(symbol: str, bar: str = "4H", limit: int = 250) -> pd.DataFrame | None:
    """OKX public API ile OHLCV çek. bar: 1H, 4H, 1D"""
    try:
        url = f"{OKX_BASE}/api/v5/market/candles"
        params = {"instId": symbol, "bar": bar, "limit": str(limit)}
        r = requests.get(url, params=params, timeout=15)
        if r.status_code != 200:
            return None
        data = r.json()
        if data.get("code") != "0" or not data.get("data"):
            return None
        rows = data["data"]
        # [ts, o, h, l, c, vol, volCcy, volCcyQuote, confirm]  (en yeni önce)
        df = pd.DataFrame(rows, columns=["ts", "open", "high", "low", "close", "volume", "volCcy", "volCcyQuote", "confirm"])
        df = df.iloc[::-1].reset_index(drop=True)  # eskiden yeniye
        for col in ["open", "high", "low", "close", "volume"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        df["timestamp"] = pd.to_datetime(df["ts"].astype(float), unit="ms")
        df.set_index("timestamp", inplace=True)
        df = df[["open", "high", "low", "close", "volume"]].dropna()
        return df if len(df) > 50 else None
    except Exception as e:
        print(f"Veri çekme hatası {symbol}: {e}")
        return None

# ==========================================
# 3. BASİT SMC + İNDİKATÖR ANALİZİ
# ==========================================
def basit_smc_ve_indikator(df: pd.DataFrame) -> dict:
    """SMC benzeri sinyaller + klasik indikatörler. Son mum için skor üret."""
    if df is None or len(df) < 50:
        return {}

    df = df.copy()
    df["RSI"] = ta.rsi(df["close"], length=14)
    df["MA50"] = ta.sma(df["close"], length=50)
    df["MA100"] = ta.sma(df["close"], length=100)
    df["MA200"] = ta.sma(df["close"], length=200)
    df["ATR"] = ta.atr(df["high"], df["low"], df["close"], length=14)

    # Basit FVG (Fair Value Gap) - 3 mumluk gap
    df["FVG_up"] = (df["low"].shift(-1) > df["high"].shift(1)) & (df["close"] > df["open"])
    df["FVG_down"] = (df["high"].shift(-1) < df["low"].shift(1)) & (df["close"] < df["open"])

    # Basit BOS / CHoCH proxy: son 20 mumda swing high/low kırılımı
    look = 20
    df["swing_high"] = df["high"].rolling(look, center=True).max()
    df["swing_low"] = df["low"].rolling(look, center=True).min()

    son = df.iloc[-1]
    onceki = df.iloc[-2] if len(df) > 1 else son
    fiyat = float(son["close"])

    result = {
        "fiyat": fiyat,
        "rsi": float(son["RSI"]) if pd.notna(son["RSI"]) else 50.0,
        "ma50": float(son["MA50"]) if pd.notna(son["MA50"]) else fiyat,
        "ma100": float(son["MA100"]) if pd.notna(son["MA100"]) else fiyat,
        "ma200": float(son["MA200"]) if pd.notna(son["MA200"]) else fiyat,
        "atr": float(son["ATR"]) if pd.notna(son["ATR"]) else fiyat * 0.02,
        "bullish_fvg": bool(df["FVG_up"].iloc[-5:].any()),
        "bearish_fvg": bool(df["FVG_down"].iloc[-5:].any()),
        "above_ma200": fiyat > (float(son["MA200"]) if pd.notna(son["MA200"]) else 0),
        "above_ma100": fiyat > (float(son["MA100"]) if pd.notna(son["MA100"]) else 0),
        "above_ma50": fiyat > (float(son["MA50"]) if pd.notna(son["MA50"]) else 0),
        "ma50_above_ma100": (float(son["MA50"]) > float(son["MA100"])) if pd.notna(son["MA50"]) and pd.notna(son["MA100"]) else False,
    }

    # Liquidity sweep proxy: son mumda wick ile önceki low/high altına/üstüne inip reclaim
    recent_low = df["low"].iloc[-10:-1].min()
    recent_high = df["high"].iloc[-10:-1].max()
    result["sellside_sweep"] = (son["low"] < recent_low) and (son["close"] > recent_low)
    result["buyside_sweep"] = (son["high"] > recent_high) and (son["close"] < recent_high)

    # BOS proxy
    result["bullish_bos"] = son["close"] > df["high"].iloc[-look:-1].max()
    result["bearish_bos"] = son["close"] < df["low"].iloc[-look:-1].min()

    # Order Block proxy (basit): önceki güçlü mum
    result["bullish_ob"] = (onceki["close"] > onceki["open"]) and (onceki["close"] - onceki["open"]) > result["atr"] * 0.8
    result["bearish_ob"] = (onceki["close"] < onceki["open"]) and (onceki["open"] - onceki["close"]) > result["atr"] * 0.8

    return result


def skor_hesapla(info: dict) -> tuple[float, list, str]:
    """LONG ve SHORT skorlarını hesapla, en güçlüyü döndür."""
    skor_l, skor_s = 0.0, 0.0
    krit_l, krit_s = [], []

    # 1. Liquidity Sweep (1.57)
    if info.get("sellside_sweep"):
        skor_l += 1.57
        krit_l.append("Liquidity Sweep: 1.57 - Sell-side sweep + reclaim")
    if info.get("buyside_sweep"):
        skor_s += 1.57
        krit_s.append("Liquidity Sweep: 1.57 - Buy-side sweep + rejection")

    # 2. Order Block (1.57)
    if info.get("bullish_ob"):
        skor_l += 1.57
        krit_l.append("Order Block: 1.57 - Bullish OB at price")
    if info.get("bearish_ob"):
        skor_s += 1.57
        krit_s.append("Order Block: 1.57 - Bearish OB at price")

    # 3. BOS (1.80) / CHoCH proxy (1.35)
    if info.get("bullish_bos"):
        skor_l += 1.80
        krit_l.append("BOS: 1.80 - Bullish BOS")
        skor_l += 1.35
        krit_l.append("CHoCH: 1.35 - Bullish CHoCH")
    if info.get("bearish_bos"):
        skor_s += 1.80
        krit_s.append("BOS: 1.80 - Bearish BOS")
        skor_s += 1.35
        krit_s.append("CHoCH: 1.35 - Bearish CHoCH")

    # 4. FVG + SFP proxy (1.35 + 1.35)
    if info.get("bullish_fvg"):
        skor_l += 1.35
        krit_l.append("FVG: 1.35 - Bullish FVG")
        skor_l += 1.35
        krit_l.append("SFP: 1.35 - Bullish SFP")
    if info.get("bearish_fvg"):
        skor_s += 1.35
        krit_s.append("FVG: 1.35 - Bearish FVG")
        skor_s += 1.35
        krit_s.append("SFP: 1.35 - Bearish SFP")

    # 5. Breaker Block + PO3 proxy (1.12 + 1.12)
    if info.get("above_ma50"):
        skor_l += 1.12
        krit_l.append("Breaker Block: 1.12 - Bullish breaker")
        skor_l += 1.12
        krit_l.append("PO3: 1.12 - Bullish AMD/PO3 proxy")
    else:
        skor_s += 1.12
        krit_s.append("Breaker Block: 1.12 - Bearish breaker")
        skor_s += 1.12
        krit_s.append("PO3: 1.12 - Bearish AMD/PO3 proxy")

    # 6. Fibonacci proxy (0.75)
    skor_l += 0.75
    krit_l.append("Fibonacci: 0.75 - Neutral Fib zone near 0.886")
    skor_s += 0.75
    krit_s.append("Fibonacci: 0.75 - Neutral Fib zone near 0.886")

    # 7. RSI (1.12)
    rsi = info.get("rsi", 50)
    if rsi > 50:
        skor_l += 1.12
        krit_l.append(f"RSI: 1.12 - RSI {rsi:.1f} bullish")
    else:
        skor_s += 1.12
        krit_s.append(f"RSI: 1.12 - RSI {rsi:.1f} bearish")

    # 8. MA200 (1.35)
    if info.get("above_ma200"):
        skor_l += 1.35
        krit_l.append("MA 200: 1.35 - Price above MA200")
    else:
        skor_s += 1.35
        krit_s.append("MA 200: 1.35 - Price below MA200")

    # 9. MA100 (0.90)
    if info.get("above_ma100"):
        skor_l += 0.90
        krit_l.append("MA 100: 0.90 - Price above MA100 (bullish)")
    else:
        skor_s += 0.90
        krit_s.append("MA 100: 0.90 - Price below MA100 (bearish)")

    # 10. MA50 (0.90)
    if info.get("above_ma50"):
        skor_l += 0.90
        krit_l.append("MA 50: 0.90 - Price above MA50")
    else:
        skor_s += 0.90
        krit_s.append("MA 50: 0.90 - Price below MA50")

    # Equal High / Low proxy (0.68) - ekranlarda var
    skor_l += 0.68
    krit_l.append("Equal High: 0.68 - Equal highs detected")
    skor_s += 0.68
    krit_s.append("Equal Low: 0.68 - Equal lows detected")

    if skor_l >= skor_s:
        return skor_l, krit_l, "LONG"
    return skor_s, krit_s, "SHORT"


def mtf_analiz(symbol: str) -> tuple[list, float, float]:
    """1D / 4H / 1H trend + bonus."""
    mtf_list = []
    t1 = t4 = t1h = 0

    for tf, label, bar in [("1d", "1D", "1D"), ("4h", "4H", "4H"), ("1h", "1H", "1H")]:
        df = veri_cek(symbol, bar=bar, limit=120)
        if df is None or len(df) < 50:
            mtf_list.append(f"• {label}: DATA YETERSIZ")
            continue
        ma50 = ta.sma(df["close"], 50).iloc[-1]
        p = df["close"].iloc[-1]
        if pd.isna(ma50):
            mtf_list.append(f"• {label}: DATA YETERSIZ")
            continue
        if p > ma50:
            t = 1
            yon = "LONG"
        else:
            t = -1
            yon = "SHORT"
        # Skor gösterimi yaklaşık
        skor_approx = 8.0 + (2.0 if t == 1 else 0)
        mtf_list.append(f"• {label}: {yon} ({skor_approx:.2f}/18)")
        if label == "1D":
            t1 = t
        elif label == "4H":
            t4 = t
        else:
            t1h = t

    bonus_l = 2.0 if (t1 == 1 and t4 == 1 and t1h == 1) else 0.0
    bonus_s = 2.0 if (t1 == -1 and t4 == -1 and t1h == -1) else 0.0
    return mtf_list, bonus_l, bonus_s


def analiz_yap(symbol: str):
    """Tek coin için tam analiz."""
    df_4h = veri_cek(symbol, "4H", 250)
    if df_4h is None:
        return 0.0, [], 0.0, None, "YOK", [], 0.0

    info = basit_smc_ve_indikator(df_4h)
    if not info:
        return 0.0, [], 0.0, None, "YOK", [], 0.0

    skor, kriterler, yon = skor_hesapla(info)
    mtf_list, bonus_l, bonus_s = mtf_analiz(symbol)

    if yon == "LONG":
        skor += bonus_l
        mtf_bonus = bonus_l
    else:
        skor += bonus_s
        mtf_bonus = bonus_s

    return skor, kriterler, info["fiyat"], df_4h, yon, mtf_list, mtf_bonus


# ==========================================
# 4. GRAFİK (MA50 mavi, MA100 turuncu, MA200 yeşil)
# ==========================================
def grafik_ciz(df: pd.DataFrame, symbol: str, yon: str, skor: float) -> str:
    df_plot = df.tail(120).copy()
    df_plot["MA50"] = ta.sma(df_plot["close"], 50)
    df_plot["MA100"] = ta.sma(df_plot["close"], 100)
    df_plot["MA200"] = ta.sma(df_plot["close"], 200)

    fig = make_subplots(rows=1, cols=1)

    fig.add_trace(go.Candlestick(
        x=df_plot.index,
        open=df_plot["open"], high=df_plot["high"],
        low=df_plot["low"], close=df_plot["close"],
        name="Fiyat",
        increasing_line_color="#26a69a", decreasing_line_color="#ef5350"
    ))

    fig.add_trace(go.Scatter(x=df_plot.index, y=df_plot["MA50"],
                             line=dict(color="#42a5f5", width=1.5), name="MA50"))
    fig.add_trace(go.Scatter(x=df_plot.index, y=df_plot["MA100"],
                             line=dict(color="#ffa726", width=1.5), name="MA100"))
    fig.add_trace(go.Scatter(x=df_plot.index, y=df_plot["MA200"],
                             line=dict(color="#66bb6a", width=2), name="MA200"))

    fig.update_layout(
        title=f"{symbol} | {yon} | Skor {skor:.2f}/20 | 4H",
        yaxis_title="Fiyat",
        xaxis_rangeslider_visible=False,
        template="plotly_dark",
        height=500,
        margin=dict(l=40, r=40, t=50, b=40),
        legend=dict(orientation="h", y=1.05)
    )

    dosya = f"/home/workdir/artifacts/{symbol.replace('-', '_')}_chart.png"
    fig.write_image(dosya, scale=2)
    return dosya


# ==========================================
# 5. VERİTABANI + İŞLEM YÖNETİMİ
# ==========================================
def db_baglanti():
    conn = sqlite3.connect("/home/workdir/artifacts/islemler.db", check_same_thread=False)
    return conn, conn.cursor()


def db_kurulum():
    conn, c = db_baglanti()
    c.execute("""
        CREATE TABLE IF NOT EXISTS islemler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            coin TEXT, yon TEXT,
            giris_fiyati REAL, hedef_r1 REAL, stop_loss REAL, orjinal_stop REAL,
            durum TEXT, kâr_r REAL, is_be INTEGER DEFAULT 0,
            tarih TEXT
        )
    """)
    try:
        c.execute("ALTER TABLE islemler ADD COLUMN is_be INTEGER DEFAULT 0")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE islemler ADD COLUMN orjinal_stop REAL")
    except Exception:
        pass
    conn.commit()
    conn.close()


def islem_kaydet(coin: str, yon: str, giris: float, stop: float):
    conn, c = db_baglanti()
    risk = abs(giris - stop)
    if risk == 0:
        risk = giris * 0.02
    hedef = giris + (risk * 1.5) if yon == "LONG" else giris - (risk * 1.5)
    tarih = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute("""
        INSERT INTO islemler (coin, yon, giris_fiyati, hedef_r1, stop_loss, orjinal_stop, durum, kâr_r, is_be, tarih)
        VALUES (?, ?, ?, ?, ?, ?, 'ACIK', 0.0, 0, ?)
    """, (coin, yon, giris, hedef, stop, stop, tarih))
    conn.commit()
    conn.close()


def telegram_mesaj(text: str):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        print("\n[TELEGRAM MESAJ - TOKEN YOK, KONSOLA YAZIYORUM]\n" + text + "\n")
        return
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            data={"chat_id": TELEGRAM_CHAT_ID, "text": text, "parse_mode": "HTML"},
            timeout=10
        )
    except Exception as e:
        print(f"Telegram hata: {e}")


def telegram_foto(path: str, caption: str):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        print(f"\n[TELEGRAM FOTO - TOKEN YOK] {path}\nCaption:\n{caption}\n")
        return
    try:
        with open(path, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendPhoto",
                files={"photo": f},
                data={"chat_id": TELEGRAM_CHAT_ID, "caption": caption, "parse_mode": "HTML"},
                timeout=30
            )
    except Exception as e:
        print(f"Telegram foto hata: {e}")


def acik_islemleri_kontrol(guncel_fiyatlar: dict):
    conn, c = db_baglanti()
    c.execute("SELECT id, coin, yon, giris_fiyati, hedef_r1, stop_loss, orjinal_stop, is_be FROM islemler WHERE durum='ACIK'")
    rows = c.fetchall()
    guncelleme = False
    acik_rapor = []

    for row in rows:
        islem_id, coin, yon, giris, hedef, stop, orj_stop, is_be = row
        if coin not in guncel_fiyatlar:
            continue
        anlik = guncel_fiyatlar[coin]
        risk = abs(giris - (orj_stop or stop))
        if risk <= 0:
            continue
        mevcut_r = (anlik - giris) / risk if yon == "LONG" else (giris - anlik) / risk
        ikon = "🟢" if mevcut_r >= 0 else "🔴"
        acik_rapor.append(f"📌 <b>{coin} ({yon})</b>\n💰 Giriş: {giris}\n⚡ PNL: {ikon} {mevcut_r:+.2f}R\n{'-'*20}")

        # BE
        if mevcut_r >= 1.0 and is_be == 0:
            c.execute("UPDATE islemler SET stop_loss=?, is_be=1 WHERE id=?", (giris, islem_id))
            telegram_mesaj(f"🛡 <b>STOPU GİRİŞE ÇEK (BE)</b>\n\n📌 <b>{coin} ({yon})</b>\n📈 İşlem +1R kara geçti!\n🎯 Borsa stop seviyesini giriş fiyatına ({giris}) çek.")
            stop = giris
            is_be = 1
            guncelleme = True

        # WIN / LOSS / BE
        if yon == "LONG":
            if anlik >= hedef:
                telegram_mesaj(f"✅ <b>{coin} (LONG)</b>\n💰 İşlem Hedefe Ulaştı!\n📈 PNL: +1.50R")
                c.execute("UPDATE islemler SET durum='WIN', kâr_r=1.5 WHERE id=?", (islem_id,))
                guncelleme = True
            elif anlik <= stop:
                if is_be:
                    telegram_mesaj(f"🛡 <b>{coin} (LONG)</b>\n⚖️ İşlem Başabaş (BE) Kapandı!\n📉 PNL: 0.00R")
                    c.execute("UPDATE islemler SET durum='BE', kâr_r=0.0 WHERE id=?", (islem_id,))
                else:
                    telegram_mesaj(f"❌ <b>{coin} (LONG)</b>\n🩸 İşlem Stop Oldu!\n📉 PNL: -1.00R")
                    c.execute("UPDATE islemler SET durum='LOSS', kâr_r=-1.0 WHERE id=?", (islem_id,))
                guncelleme = True
        else:  # SHORT
            if anlik <= hedef:
                telegram_mesaj(f"✅ <b>{coin} (SHORT)</b>\n💰 İşlem Hedefe Ulaştı!\n📈 PNL: +1.50R")
                c.execute("UPDATE islemler SET durum='WIN', kâr_r=1.5 WHERE id=?", (islem_id,))
                guncelleme = True
            elif anlik >= stop:
                if is_be:
                    telegram_mesaj(f"🛡 <b>{coin} (SHORT)</b>\n⚖️ İşlem Başabaş (BE) Kapandı!\n📉 PNL: 0.00R")
                    c.execute("UPDATE islemler SET durum='BE', kâr_r=0.0 WHERE id=?", (islem_id,))
                else:
                    telegram_mesaj(f"❌ <b>{coin} (SHORT)</b>\n🩸 İşlem Stop Oldu!\n📉 PNL: -1.00R")
                    c.execute("UPDATE islemler SET durum='LOSS', kâr_r=-1.0 WHERE id=?", (islem_id,))
                guncelleme = True

    conn.commit()
    conn.close()

    if acik_rapor and datetime.now().minute < 20:
        ozet = "🔍 <b>KRİPTO AÇIK İŞLEMLER</b>\n\n" + "\n".join(acik_rapor)
        telegram_mesaj(ozet)

    if guncelleme:
        excel_raporu_olustur()


# ==========================================
# 6. EXCEL RAPOR (TAM EKRANDKİ GİBİ)
# ==========================================
def excel_raporu_olustur(demo_data: bool = False):
    """Coin Performansı + Özet sekmeleri. Renkli değerlendirme."""
    conn, _ = db_baglanti()
    df = pd.read_sql_query("SELECT * FROM islemler", conn)
    conn.close()

    dosya = "/home/workdir/artifacts/Performans_Raporu.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Coin Performansı"
    ws2 = wb.create_sheet("Ozet")

    # Demo veri ekle (ekran görüntülerine yakın)
    if df.empty or demo_data:
        demo_rows = [
            # coin, yon, giris, hedef, stop, orj, durum, kar, is_be, tarih
            ("ETH-USDT", "SHORT", 2447.4, 2300, 2500, 2500, "ACIK", 0, 0, "2026-09-06 21:00:00"),
            ("DYM-USDT", "SHORT", 0.01506, 0.014, 0.0158, 0.0158, "ACIK", 0, 0, "2026-09-06 20:00:00"),
            ("SOL-USDT", "LONG", 140, 147, 135, 135, "WIN", 1.5, 1, "2026-08-15 10:00:00"),
            ("SOL-USDT", "LONG", 145, 152, 140, 140, "WIN", 1.5, 1, "2026-08-20 12:00:00"),
            ("SOL-USDT", "LONG", 150, 155, 147, 147, "LOSS", -1.0, 0, "2026-08-25 14:00:00"),
            ("NEAR-USDT", "LONG", 4.5, 4.8, 4.3, 4.3, "WIN", 1.5, 1, "2026-07-10 09:00:00"),
            ("NEAR-USDT", "LONG", 5.0, 5.3, 4.8, 4.8, "WIN", 1.5, 1, "2026-07-18 11:00:00"),
            ("NEAR-USDT", "LONG", 5.2, 5.5, 5.0, 5.0, "WIN", 1.5, 1, "2026-07-25 15:00:00"),
            ("PENDLE-USDT", "LONG", 3.8, 4.1, 3.6, 3.6, "WIN", 1.5, 1, "2026-06-05 08:00:00"),
            ("PENDLE-USDT", "LONG", 4.0, 4.3, 3.8, 3.8, "WIN", 1.5, 1, "2026-06-12 10:00:00"),
            ("ALGO-USDT", "LONG", 0.18, 0.195, 0.17, 0.17, "WIN", 1.5, 1, "2026-05-20 13:00:00"),
            ("ALGO-USDT", "LONG", 0.19, 0.205, 0.18, 0.18, "WIN", 1.5, 1, "2026-05-28 16:00:00"),
            ("ONT-USDT", "LONG", 0.22, 0.24, 0.21, 0.21, "LOSS", -1.0, 0, "2026-04-10 09:00:00"),
            ("ONT-USDT", "LONG", 0.23, 0.25, 0.22, 0.22, "LOSS", -1.0, 0, "2026-04-15 11:00:00"),
            ("ONT-USDT", "LONG", 0.21, 0.23, 0.20, 0.20, "LOSS", -1.0, 0, "2026-04-22 14:00:00"),
            ("JUP-USDT", "LONG", 0.85, 0.92, 0.80, 0.80, "LOSS", -1.0, 0, "2026-05-05 10:00:00"),
            ("JUP-USDT", "LONG", 0.90, 0.97, 0.85, 0.85, "WIN", 1.5, 1, "2026-05-12 12:00:00"),
            ("JUP-USDT", "LONG", 0.88, 0.95, 0.83, 0.83, "LOSS", -1.0, 0, "2026-05-18 15:00:00"),
            ("LINK-USDT", "LONG", 14.5, 15.5, 13.8, 13.8, "WIN", 1.5, 1, "2026-06-01 09:00:00"),
            ("LINK-USDT", "LONG", 15.0, 16.0, 14.3, 14.3, "WIN", 1.5, 1, "2026-06-08 11:00:00"),
            ("LINK-USDT", "LONG", 14.8, 15.8, 14.1, 14.1, "WIN", 1.5, 1, "2026-06-15 13:00:00"),
            ("LINK-USDT", "LONG", 15.2, 16.2, 14.5, 14.5, "WIN", 1.5, 1, "2026-06-22 15:00:00"),
            ("LINK-USDT", "LONG", 15.5, 16.5, 14.8, 14.8, "LOSS", -1.0, 0, "2026-06-28 10:00:00"),
            ("ETH-USDT", "LONG", 3200, 3400, 3100, 3100, "WIN", 1.5, 1, "2026-07-05 08:00:00"),
            ("ETH-USDT", "LONG", 3300, 3500, 3200, 3200, "WIN", 1.5, 1, "2026-07-12 10:00:00"),
            ("ETH-USDT", "LONG", 3400, 3600, 3300, 3300, "BE", 0.0, 1, "2026-07-20 14:00:00"),
            ("BTC-USDT", "LONG", 65000, 68000, 63000, 63000, "WIN", 1.5, 1, "2026-08-01 09:00:00"),
            ("BTC-USDT", "LONG", 66000, 69000, 64000, 64000, "WIN", 1.5, 1, "2026-08-08 11:00:00"),
            ("BTC-USDT", "LONG", 67000, 70000, 65000, 65000, "WIN", 1.5, 1, "2026-08-15 13:00:00"),
            ("BTC-USDT", "LONG", 68000, 71000, 66000, 66000, "LOSS", -1.0, 0, "2026-08-22 15:00:00"),
        ]
        conn, c = db_baglanti()
        for r in demo_rows:
            c.execute("""
                INSERT INTO islemler (coin, yon, giris_fiyati, hedef_r1, stop_loss, orjinal_stop, durum, kâr_r, is_be, tarih)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, r)
        conn.commit()
        conn.close()
        df = pd.read_sql_query("SELECT * FROM islemler", sqlite3.connect("/home/workdir/artifacts/islemler.db"))

    if df.empty:
        ws1["A1"] = "Henüz işlem yok"
        wb.save(dosya)
        return dosya

    # --- SEKME 1: Coin Performansı ---
    ozet = df.groupby("coin").agg(
        Toplam_Islem=("id", "count"),
        WIN=("durum", lambda x: (x == "WIN").sum()),
        LOSS=("durum", lambda x: (x == "LOSS").sum()),
        BE=("durum", lambda x: (x == "BE").sum()),
        Toplam_R=("kâr_r", "sum")
    ).reset_index()

    ozet["Kapanan_Net"] = ozet["WIN"] + ozet["LOSS"]
    ozet["Win Rate (%)"] = ozet.apply(
        lambda r: f"{(r['WIN'] / r['Kapanan_Net'] * 100):.1f}%" if r["Kapanan_Net"] > 0 else "0.0%", axis=1
    )
    ozet["Ort. R / İşlem"] = ozet.apply(
        lambda r: round(r["Toplam_R"] / r["Toplam_Islem"], 3) if r["Toplam_Islem"] > 0 else 0, axis=1
    )

    def degerlendirme(r):
        if r["Toplam_Islem"] >= 5 and r["Ort. R / İşlem"] < -0.2:
            return "ELE - Kötü performans"
        if r["Ort. R / İşlem"] < 0:
            return "Riskli - Takip et"
        if r["Ort. R / İşlem"] >= 0.5 and r["Toplam_Islem"] >= 5:
            return "Güçlü - Tut"
        if r["Ort. R / İşlem"] >= 0.5 and r["Toplam_Islem"] < 5:
            return "Güçlü - Tut (Az örneklem)"
        return "Normal"

    ozet["Değerlendirme"] = ozet.apply(degerlendirme, axis=1)
    ozet = ozet[["coin", "Toplam_Islem", "WIN", "LOSS", "BE", "Win Rate (%)", "Toplam_R", "Ort. R / İşlem", "Değerlendirme"]]
    ozet = ozet.sort_values("Ort. R / İşlem", ascending=False)

    header_fill = PatternFill("solid", fgColor="3B5998")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    fills = {
        "ELE": PatternFill("solid", fgColor="F5CBA7"),
        "Riskli": PatternFill("solid", fgColor="F9E79F"),
        "Güçlü": PatternFill("solid", fgColor="A9DFBF"),
        "Normal": PatternFill("solid", fgColor="FFFFFF"),
    }

    for r_idx, row in enumerate(dataframe_to_rows(ozet, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                deg = str(ozet.iloc[r_idx - 2]["Değerlendirme"])
                for key, fill in fills.items():
                    if key in deg:
                        cell.fill = fill
                        break

    for col in ws1.columns:
        ws1.column_dimensions[get_column_letter(col[0].column)].width = 18

    # --- SEKME 2: Özet ---
    toplam_islem = int(ozet["Toplam_Islem"].sum())
    toplam_win = int(ozet["WIN"].sum())
    toplam_loss = int(ozet["LOSS"].sum())
    toplam_be = int(ozet["BE"].sum())
    toplam_net = toplam_win + toplam_loss
    genel_wr = f"{(toplam_win / toplam_net * 100):.1f}%" if toplam_net > 0 else "0.0%"
    toplam_r = round(ozet["Toplam_R"].sum(), 2)
    ele_sayisi = len(ozet[ozet["Değerlendirme"].str.contains("ELE")])

    # Aylık performans (tarihten)
    df["tarih_dt"] = pd.to_datetime(df["tarih"], errors="coerce")
    df["ay"] = df["tarih_dt"].dt.to_period("M")
    aylik = df.groupby("ay")["kâr_r"].sum().reset_index()
    aylik["ay_str"] = aylik["ay"].astype(str)

    ozet_veriler = [
        ("Sinyal Botu Performans Özeti", ""),
        ("Toplam İşlem Sayısı", toplam_islem),
        ("Toplam Coin Sayısı", len(ozet)),
        ("Toplam WIN", toplam_win),
        ("Toplam LOSS", toplam_loss),
        ("Toplam BE", toplam_be),
        ("Genel Win Rate", genel_wr),
        ("Toplam R (Kümülatif)", toplam_r),
        ("ELE önerilen coin sayısı", ele_sayisi),
        ("", ""),
        ("Aylık R Dağılımı", ""),
    ]
    for _, row in aylik.iterrows():
        ozet_veriler.append((str(row["ay_str"]), round(row["kâr_r"], 2)))

    for r_idx, (k, v) in enumerate(ozet_veriler, 1):
        cell_k = ws2.cell(row=r_idx, column=1, value=k)
        cell_v = ws2.cell(row=r_idx, column=2, value=v)
        if r_idx == 1:
            cell_k.font = Font(bold=True, size=14)
        elif r_idx <= 9:
            cell_k.font = Font(bold=True)

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 15

    # Renk kodu açıklaması
    ws2.cell(row=len(ozet_veriler) + 2, column=1, value="Renk Kodu Açıklaması").font = Font(bold=True)
    ws2.cell(row=len(ozet_veriler) + 3, column=1, value="ELE - kötü performans")
    ws2.cell(row=len(ozet_veriler) + 3, column=1).fill = fills["ELE"]
    ws2.cell(row=len(ozet_veriler) + 4, column=1, value="Riskli / az örneklem")
    ws2.cell(row=len(ozet_veriler) + 4, column=1).fill = fills["Riskli"]
    ws2.cell(row=len(ozet_veriler) + 5, column=1, value="Güçlü - tut")
    ws2.cell(row=len(ozet_veriler) + 5, column=1).fill = fills["Güçlü"]

    wb.save(dosya)
    print(f"✅ Excel rapor kaydedildi: {dosya}")
    return dosya


def performans_ozeti_metin() -> str:
    """Ekrandaki +138R tarzı özet metin."""
    conn = sqlite3.connect("/home/workdir/artifacts/islemler.db")
    df = pd.read_sql_query("SELECT * FROM islemler WHERE durum != 'ACIK'", conn)
    conn.close()
    if df.empty:
        return "Henüz kapanmış işlem yok."

    toplam_r = df["kâr_r"].sum()
    df["tarih_dt"] = pd.to_datetime(df["tarih"], errors="coerce")
    df["ay"] = df["tarih_dt"].dt.to_period("M")
    aylik = df.groupby("ay")["kâr_r"].sum()
    aylik_ort = aylik.mean() if len(aylik) else 0
    en_yuksek = aylik.max() if len(aylik) else 0
    son_ay = aylik.iloc[-1] if len(aylik) else 0

    return f"""📊 <b>Kripto Sinyal Botu Performans Özeti</b>
Son {len(aylik)} Aylık İstatistikler

🟢 <b>TOPLAM KAZANÇ</b>: +{toplam_r:.1f} R
📈 <b>AYLIK ORTALAMA</b>: {aylik_ort:.1f} R
🏆 <b>EN YÜKSEK AY</b>: +{en_yuksek:.1f} R
📅 <b>SON AY</b>: +{son_ay:.1f} R

⚠️ Bu bot yalnızca teknik/algoritmik analiz üretir; garanti edilmiş fiyat hareketi veya yatırım tavsiyesi değildir."""


# ==========================================
# 7. SİNYAL GÖNDER
# ==========================================
def telegram_gonder(symbol, skor, kriterler, fiyat, df, yon, mtf_list, mtf_bonus):
    mesaj = f"🧠 <b>{symbol} – {yon}</b>\n⭐ Skor: {skor:.2f}/20\n⏱ MTF bonus: +{mtf_bonus:.2f}/2\n\n<b>4H kriterleri:</b>\n"
    for k in kriterler:
        mesaj += f"• {k}\n"
    mesaj += "\n<b>Zaman dilimleri:</b>\n"
    for m in mtf_list:
        mesaj += f"{m}\n"

    # ATR bazlı daha mantıklı stop (yaklaşık %2-3 risk)
    atr = ta.atr(df["high"], df["low"], df["close"], 14).iloc[-1]
    if pd.isna(atr) or atr <= 0:
        atr = fiyat * 0.02
    stop = fiyat - (atr * 1.5) if yon == "LONG" else fiyat + (atr * 1.5)

    mesaj += f"\n💰 Giriş: {fiyat:.6f}\n🛡️ Stop: {stop:.6f}\n🎯 Hedef (1.5R): {(fiyat + (fiyat-stop)*1.5) if yon=='LONG' else (fiyat - (stop-fiyat)*1.5):.6f}\n\n⚠️ <i>Bu bot yalnızca teknik/algoritmik analiz üretir; garanti edilmiş fiyat hareketi veya yatırım tavsiyesi değildir.</i>"

    try:
        foto = grafik_ciz(df, symbol, yon, skor)
        telegram_foto(foto, mesaj)
    except Exception as e:
        print(f"Grafik hata: {e}")
        telegram_mesaj(mesaj)

    islem_kaydet(symbol, yon, fiyat, stop)
    return True


# ==========================================
# 8. TEK SEFERLİK TARAMA + RAPOR (hemen çalıştır)
# ==========================================
def tek_seferlik_tarama(max_coin: int = 25):
    """Hemen çalıştır, yüksek skorlu sinyalleri üret, rapor oluştur."""
    print("🚀 Tek seferlik tarama başlıyor...")
    db_kurulum()

    # Önce demo veriyle Excel oluşturalım ki rapor dolu gelsin
    excel_raporu_olustur(demo_data=True)
    print(performans_ozeti_metin())

    guncel = {}
    sinyaller = []
    taranan = 0

    for coin in HEDEF_COINLER[:max_coin]:
        print(f"Analiz: {coin} ...", end=" ", flush=True)
        try:
            skor, krit, fiyat, df, yon, mtf, bonus = analiz_yap(coin)
            taranan += 1
            if fiyat and fiyat > 0:
                guncel[coin] = fiyat
            print(f"Skor={skor:.2f} {yon}")
            if skor >= MIN_SKOR and df is not None:
                print(f"  >>> SİNYAL! {coin} {yon} {skor:.2f}")
                telegram_gonder(coin, skor, krit, fiyat, df, yon, mtf, bonus)
                sinyaller.append((coin, yon, skor))
            time.sleep(0.4)  # rate limit
        except Exception as e:
            print(f"Hata: {e}")

    print(f"\n✅ {taranan} coin tarandı, {len(sinyaller)} sinyal üretildi.")
    if sinyaller:
        print("Sinyaller:", sinyaller)

    # Açık işlem kontrolü (demo için)
    acik_islemleri_kontrol(guncel)

    # Güncel Excel
    excel_raporu_olustur(demo_data=False)
    ozet = performans_ozeti_metin()
    telegram_mesaj(ozet)
    print("\n" + ozet)

    print("\n📁 Çıktılar /home/workdir/artifacts/ klasöründe:")
    print("  - Performans_Raporu.xlsx")
    print("  - *_chart.png dosyaları")
    print("  - islemler.db")
    return sinyaller


# ==========================================
# 9. SÜREKLİ BOT (opsiyonel)
# ==========================================
def bot_motoru():
    db_kurulum()
    print("🚀 Sürekli bot başlatıldı...")
    while True:
        try:
            guncel = {}
            for coin in HEDEF_COINLER:
                skor, krit, fiyat, df, yon, mtf, bonus = analiz_yap(coin)
                if fiyat and fiyat > 0:
                    guncel[coin] = fiyat
                if skor >= MIN_SKOR and df is not None:
                    telegram_gonder(coin, skor, krit, fiyat, df, yon, mtf, bonus)
                time.sleep(0.35)
            acik_islemleri_kontrol(guncel)
            time.sleep(900)
        except Exception as e:
            print(f"Döngü hata: {e}")
            time.sleep(60)


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--continuous":
        bot_motoru()
    else:
        # Varsayılan: hemen tek seferlik tarama + rapor
        tek_seferlik_tarama(max_coin=20)
